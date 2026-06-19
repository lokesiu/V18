"""
xlsx_renderer.py - Excel Document Generator for 明证台 V18

Generates the 04_证据目录.xlsx evidence catalog spreadsheet.
"""
from __future__ import annotations

import os
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from core.fact_card import FactCard

logger = logging.getLogger(__name__)

# Style constants
HEADER_FONT = Font(name="黑体", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="宋体", size=10)
BODY_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Column definitions: (header, width)
COLUMNS = [
    ("序号", 8),
    ("证据名称", 30),
    ("证据类型", 12),
    ("证明内容", 35),
    ("来源", 20),
    ("页码", 10),
    ("备注", 20),
]


def render_evidence_catalog(fact_card: FactCard, output_path: str) -> bool:
    """Render 04_证据目录.xlsx with evidence catalog.

    Columns: 序号, 证据名称, 证据类型, 证明内容, 来源, 页码, 备注
    Must have >= 3 rows of real data; pad with "待补充" if fewer.

    Args:
        fact_card: FactCard with evidence source_refs.
        output_path: Output file path.

    Returns:
        True on success, False on failure.
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "证据目录"

        # --- Set column widths ---
        for col_idx, (_, width) in enumerate(COLUMNS, 1):
            col_letter = chr(64 + col_idx)  # A, B, C, ...
            ws.column_dimensions[col_letter].width = width

        # --- Header row ---
        for col_idx, (header, _) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # Freeze header row
        ws.freeze_panes = "A2"

        # --- Build data rows ---
        rows_data: list[tuple[str, str, str, str, str, str, str]] = []

        if fact_card.source_refs:
            for ref in fact_card.source_refs:
                rows_data.append((
                    "",  # 序号 filled later
                    ref.file_name or "未知文件",
                    "书证",
                    ref.excerpt or "证明相关事实",
                    "当事人提供" if ref.file_name else "待补充",
                    str(ref.page) if ref.page else "",
                    "",
                ))

        # Pad to at least 3 rows
        while len(rows_data) < 3:
            rows_data.append((
                "",
                "待补充",
                "待补充",
                "待补充",
                "待补充",
                "",
                "待补充证据材料",
            ))

        # --- Write data rows ---
        for row_idx, row_data in enumerate(rows_data, 2):
            seq = row_idx - 1  # 序号 starts from 1
            values = [seq] + list(row_data[1:])  # Replace empty 序号 with seq
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                if col_idx in (1, 6):  # 序号 and 页码 centered
                    cell.alignment = CENTER_ALIGNMENT
                else:
                    cell.alignment = BODY_ALIGNMENT

        # --- Add summary row ---
        summary_row = len(rows_data) + 3
        ws.cell(row=summary_row, column=1, value="汇总").font = Font(
            name="黑体", size=10, bold=True
        )
        ws.cell(row=summary_row, column=2, value=f"共{len(rows_data)}份证据材料").font = BODY_FONT

        if fact_card.case_id:
            ws.cell(row=summary_row + 1, column=1, value="案件编号").font = Font(
                name="黑体", size=10, bold=True
            )
            ws.cell(row=summary_row + 1, column=2, value=fact_card.case_id).font = BODY_FONT

        # --- Save ---
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
        wb.close()

        logger.info("Evidence catalog rendered: %s", output_path)
        return True

    except Exception as exc:
        logger.error("Failed to render evidence catalog: %s", exc)
        return False


# Alias for pipeline compatibility
def render_xlsx(fact_card: FactCard, output_path: str) -> bool:
    """Render XLSX evidence catalog. Wrapper for render_evidence_catalog."""
    return render_evidence_catalog(fact_card, output_path)
