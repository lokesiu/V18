"""
package_leak_scanner.py - Quality Gate: Package Leak Scanner

Scans customer output directories for internal/debug files that should not
be delivered to customers. Also validates PDF headers and XLSX data completeness.
"""
from __future__ import annotations

import os
from typing import List

from core.quality.final_artifact_auditor import CheckResult


# File extensions that should never appear in customer output
FORBIDDEN_EXTENSIONS: List[str] = [".json", ".md"]

# File/directory names that indicate internal/debug content
FORBIDDEN_NAMES: List[str] = ["debug", "ai_raw_outputs", "raw", "internal", "trace"]


def _is_forbidden_name(name: str) -> bool:
    """Check if a file or directory name matches any forbidden pattern."""
    name_lower = name.lower()
    for forbidden in FORBIDDEN_NAMES:
        if forbidden in name_lower:
            return True
    return False


def scan_for_leaks(customer_dir: str) -> CheckResult:
    """Scan customer directory for internal/debug files that should not be there.

    Walks the entire directory tree and checks:
    - No files have forbidden extensions (.json, .md)
    - No files or directories have forbidden names (debug, ai_raw_outputs, etc.)

    Args:
        customer_dir: Absolute path to the customer output directory.

    Returns:
        CheckResult with pass/fail status and details of any violations.
    """
    if not os.path.isdir(customer_dir):
        return CheckResult(
            check_name="无泄漏文件",
            passed=False,
            message=f"目录不存在: {customer_dir}",
        )

    violations: List[str] = []

    for root, dirs, files in os.walk(customer_dir):
        # Check directory names
        for dir_name in dirs:
            if _is_forbidden_name(dir_name):
                rel_path = os.path.relpath(
                    os.path.join(root, dir_name), customer_dir
                )
                violations.append(f"禁止的目录: {rel_path}")

        # Check file names and extensions
        for file_name in files:
            # Check forbidden extensions
            _, ext = os.path.splitext(file_name)
            if ext.lower() in FORBIDDEN_EXTENSIONS:
                rel_path = os.path.relpath(
                    os.path.join(root, file_name), customer_dir
                )
                violations.append(f"禁止的文件类型({ext}): {rel_path}")

            # Check forbidden names
            if _is_forbidden_name(file_name):
                rel_path = os.path.relpath(
                    os.path.join(root, file_name), customer_dir
                )
                violations.append(f"禁止的文件名: {rel_path}")

    if violations:
        return CheckResult(
            check_name="无泄漏文件",
            passed=False,
            message=f"发现{len(violations)}个不应存在的文件/目录",
            details="\n".join(violations),
        )

    return CheckResult(
        check_name="无泄漏文件",
        passed=True,
        message="客户目录中无内部/调试文件泄漏",
    )


def check_pdf_header(pdf_path: str) -> CheckResult:
    """Verify PDF starts with %PDF header.

    Reads the first 5 bytes of the file and checks for the PDF magic bytes.

    Args:
        pdf_path: Absolute path to the .pdf file.

    Returns:
        CheckResult with pass/fail status.
    """
    basename = os.path.basename(pdf_path)

    if not os.path.isfile(pdf_path):
        return CheckResult(
            check_name="PDF文件头",
            passed=False,
            message=f"文件不存在: {basename}",
        )

    try:
        with open(pdf_path, "rb") as f:
            header = f.read(5)

        if len(header) < 4:
            return CheckResult(
                check_name="PDF文件头",
                passed=False,
                message=f"PDF文件过小({len(header)}字节): {basename}",
            )

        if not header.startswith(b"%PDF"):
            return CheckResult(
                check_name="PDF文件头",
                passed=False,
                message=f"PDF文件头无效: {basename} (实际: {header[:5]!r})",
            )

        return CheckResult(
            check_name="PDF文件头",
            passed=True,
            message=f"PDF文件头有效: {basename}",
        )

    except PermissionError:
        return CheckResult(
            check_name="PDF文件头",
            passed=False,
            message=f"无权读取PDF文件: {basename}",
        )
    except Exception as exc:
        return CheckResult(
            check_name="PDF文件头",
            passed=False,
            message=f"PDF文件头检查失败: {basename} - {exc}",
        )


def check_xlsx_rows(xlsx_path: str, min_rows: int = 3) -> CheckResult:
    """Verify XLSX has at least min_rows data rows.

    Uses openpyxl to count rows that contain at least one non-empty cell.
    The header row is excluded from the count.

    Args:
        xlsx_path: Absolute path to the .xlsx file.
        min_rows: Minimum number of data rows required (default: 3).

    Returns:
        CheckResult with pass/fail status.
    """
    basename = os.path.basename(xlsx_path)

    if not os.path.isfile(xlsx_path):
        return CheckResult(
            check_name="XLSX数据行",
            passed=False,
            message=f"文件不存在: {basename}",
        )

    try:
        from openpyxl import load_workbook
    except ImportError:
        return CheckResult(
            check_name="XLSX数据行",
            passed=False,
            message=f"openpyxl未安装，无法检查: {basename}",
        )

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as exc:
        return CheckResult(
            check_name="XLSX数据行",
            passed=False,
            message=f"无法打开XLSX: {basename} - {exc}",
        )

    total_data_rows = 0
    sheet_details: List[str] = []

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data_rows = 0
            for row in ws.iter_rows(min_row=2):  # Skip header row
                # Check if any cell in the row has content
                has_content = any(
                    cell.value is not None and str(cell.value).strip()
                    for cell in row
                )
                if has_content:
                    data_rows += 1
            total_data_rows += data_rows
            sheet_details.append(f"工作表'{sheet_name}': {data_rows}行数据")
    finally:
        wb.close()

    if total_data_rows < min_rows:
        return CheckResult(
            check_name="XLSX数据行",
            passed=False,
            message=f"XLSX数据行不足: {basename} ({total_data_rows}行, 需>={min_rows})",
            details="\n".join(sheet_details),
        )

    return CheckResult(
        check_name="XLSX数据行",
        passed=True,
        message=f"XLSX数据行充足: {basename} ({total_data_rows}行)",
    )
